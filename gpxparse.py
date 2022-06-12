#!/usr/bin/env python3
from collections import deque
from distutils import extension
from openpyxl import Workbook
from openpyxl.styles import Font, DEFAULT_FONT
from os.path import splitext
import sys, csv, argparse, textwrap
import logging,coloredlogs
from re import sub
import math, statistics
import gpxpy
import gpxpy.gpx
import geopy
from geopy.distance import geodesic
from geopy.geocoders import Nominatim

#TODO: make tracks and waypoints a class

parser = argparse.ArgumentParser(
    formatter_class=argparse.RawDescriptionHelpFormatter,
    description=textwrap.dedent('''\
        FLT Map GPX Segment Processor
        -----------------------------
          You must specify a GPX file to process and optionally a track.  If you do not specify the track you will be prompted.
          The output will be Excel, however you can specify CSV via the -c (--csv) optional command parameter.
    '''))
parser.add_argument('gpxfile',help='Specify the gpx file to process')
parser.add_argument('-c','--csv',action='store_true',help='Output as CSV instead of Excel')
parser.add_argument('-r','--reverse',action='store_true',help='Reverse the direction of the track')
parser.add_argument('-t','--track',help='Specify the track name to process')
parser.add_argument('-v','--verbose',action='count',default=0,help='Verbose mode.  Multiple -v options increase the verbosity.  The maximum is 3')
symbols = parser.add_mutually_exclusive_group()
symbols.add_argument('-i','--include',action='append',help='Symbol to include.  Use multiple -i options for additional symbols.  Cannot be used with -x/--exclude')
symbols.add_argument('-x','--exclude',action='append',help='Symbol to exclude.  Use multiple -x options for additional symbols.  Cannot be used with -i/--include')
args = vars(parser.parse_args())

# 0: log level 3 (warning)
# 1: log level 2 (info)
# 2: log level 1 (debug)
LOG_LEVEL = max((3-args.get('verbose')),1)*10

## Helper Functions ##
def initLog():
    format_string = '%(asctime)s [%(levelname)8s]:  %(message)s'
    global logger
    logger = logging.getLogger(__name__)
    logger.setLevel(LOG_LEVEL)
    levelStyles = {
        'debug':{'color':'cyan'},
        'info':{'color':'white'},
        'warning':{'color':'yellow'},
        'error':{'color':'red'},
        'critical':{'bold':True,'color':'magenta'}
    }
    coloredlogs.install(logger=logger,level=LOG_LEVEL,fmt=format_string,level_styles=levelStyles)

def camel_case(s):
  s = sub(r"[^0-9a-zA-Z]+", " ", s).title().replace(" ", "")
  return ''.join([s[0].lower(), s[1:]])

def calc_bearing(pointA:gpxpy.gpx.GPXTrackPoint,pointB:gpxpy.gpx.GPXTrackPoint) -> float:
    deg2rad = math.pi/180
    latA = pointA.latitude * deg2rad
    latB = pointB.latitude * deg2rad
    lonA = pointA.longitude * deg2rad
    lonB = pointB.longitude * deg2rad
    delta_ratio = math.log(math.tan(latB/ 2 + math.pi / 4) / math.tan(latA/ 2 + math.pi / 4))
    delta_lon = abs(lonA - lonB)

    delta_lon %= math.pi
    bearing = math.atan2(delta_lon, delta_ratio)/deg2rad
    return bearing

## Main ##
initLog();
try:
    gpx_file = open(args.get('gpxfile'),'r')
except:
    logger.critical('Could not open GPX File: {}'.format(args.get('gpxfile')))
    sys.exit(1)
gpx = gpxpy.parse(gpx_file)

# Get track to process
trkNum = None
if args.get('track') is not None:
    for i,track in enumerate(gpx.tracks):
        if track.name == args.get('track'): trkNum = i
    if trkNum is None:
        print('\nNo track with that name found, please make a selection.')

while trkNum is None:
    print('\nTrack List for {}:'.format(args.get('gpxfile')))
    for i,track in enumerate(gpx.tracks):
        print('  {0}.\t{1}'.format(int(i)+1,track.name))
    print('\n  X.\tExit')
    n = input('\nEnter the # of the track to process: ')
    if n == 'x' or n == 'X': sys.exit(0)
    if int(n) > len(gpx.tracks):
        print('Invalid track selection.  Please select a valid track #')
        trkNum = None
    else:
        trkNum = int(n)-1
        print('')

track = gpx.tracks[trkNum]
logger.info('Beginning Processing of {0} for track {1}'.format(args.get('gpxfile'),track.name))

# Collect Waypoints
logger.info('Processing waypoints...')
waypoints = deque()
wpt_symbols = dict()
for waypoint in gpx.waypoints:
    # included/exclude waypoint symbols
    if (args.get('include') is not None and waypoint.symbol in (args.get('include') or [])) or (args.get('exclude') is not None and waypoint.symbol not in (args.get('exclude') or [])):
        waypoints.append({'name':waypoint.name,'description':waypoint.description,'symbol':waypoint.symbol,'latitude':waypoint.latitude,'longitude':waypoint.longitude,'nearest_trkpoint':None,'nearest_distance':math.inf})
        wpt_symbols[camel_case(waypoint.symbol)] = waypoint.symbol
#    geolocator = Nominatim(user_agent="Finger Lakes Trail Mapper")
#    location = geolocator.reverse('{0},{1}'.format(waypoint.latitude, waypoint.longitude))
#    print(location.raw)
logger.info('Including {0} Symbols: "{1}"'.format(len(wpt_symbols),'","'.join(sorted(wpt_symbols.values()))))
if args.get('exclude') is not None: logger.info('Ignoring {0} Symbols: "{1}"'.format(len(args.get('exclude')),'","'.join(sorted(args.get('exclude')))))
logger.info('Processed {} waypoints'.format(len(waypoints)))

# Collect Track Point
logger.info('Processing trackpoints for {}...'.format(track.name))
trkpoints = deque()
pointNum = 0;
for point,s,n in track.walk():
    trkpoints.append({'number':n,'point':point,'near_waypoint':None,'distance_from_waypoint':None})
    pointNum = n
totalPoints = int(pointNum)+1
logger.info('Processed {} trackpoints'.format(totalPoints))

# reverse the track direction if option indicated
if args.get('reverse'):
    logger.warning('Reversing track direction')
    trkpoints.reverse()

# Find closest waypoint to trackpoint and update both collections
for i,waypoint in enumerate(waypoints):
    for j,trkpoint in enumerate(trkpoints):
        dist = geodesic((trkpoint['point'].latitude,trkpoint['point'].longitude),(waypoint['latitude'],waypoint['longitude'])).feet
        if dist < waypoint['nearest_distance']:
            waypoints[i]['nearest_trkpoint'] = j
            waypoints[i]['nearest_distance'] = dist
for k,waypoint in enumerate(waypoints):
    if waypoint['nearest_trkpoint'] is not None:
        trkpoints[waypoint['nearest_trkpoint']]['near_waypoint'] = k
        trkpoints[waypoint['nearest_trkpoint']]['distance_from_waypoint'] = waypoint['nearest_distance']

# Build segments
gpx_track = gpxpy.gpx.GPXTrack()
seg = deque()
#seg.append({'seg':gpxpy.gpx.GPXTrackSegment(),'wpt':{'name':'Start','description':'start'}})
#gpx_track.segments.append(seg[0]['seg'])
seg.append({'seg':gpxpy.gpx.GPXTrackSegment(),'wpt':None})
gpx_track.segments.append(seg[0]['seg'])
s = 0
for i,trkpoint in enumerate(trkpoints):
    if trkpoint['near_waypoint'] is not None:
        seg[s]['wpt'] = waypoints[trkpoint['near_waypoint']]
        s += 1
        seg.append({'seg':gpxpy.gpx.GPXTrackSegment(),'wpt':None})
        gpx_track.segments.append(seg[s]['seg'])
    seg[s]['seg'].points.append(trkpoint['point'])
seg.pop()
gpx_track.segments.pop()

runTotal = 0
rows = []
for i,segment in enumerate(gpx_track.segments):
    direction = ''
    segDist = round(geopy.units.mi(meters=segment.length_2d()),1)
    if args.get('csv'):
        runTotal += segDist
        runTotal = round(runTotal,1)
    else: 
        runTotal = '=INDIRECT(ADDRESS(ROW(),COLUMN()-1))' if i == 0 else '=INDIRECT(ADDRESS(ROW(),COLUMN()-1))+INDIRECT(ADDRESS(ROW()-1,COLUMN()))'
    if i < len(gpx_track.segments)-1:
        cs_lastp = segment.get_points_no()
        segment2 = seg[i+1]['seg']
        h0 = None if cs_lastp < 2 else calc_bearing(segment.points[cs_lastp-3],segment.points[cs_lastp-1])
        h1 = None if segment2.get_points_no() < 2 else calc_bearing(segment2.points[0],segment2.points[2])
        dH = h0 or h1
        if h0 is not None and h1 is not None: dH = ((h1-h0)+360)%360
        dirlist = ['continue straight','slight right','turn right','sharp right','u-turn to the right','turn around','u-turn to the left','sharp left','turn left','slight left','continue straight']
        if args.get('reverse'): dirlist.reverse()
        if dH >= 0 and dH <= 10: direction = dirlist[0]
        if dH > 10 and dH <= 30: direction = dirlist[1] #slight 
        if dH > 30 and dH <= 110: direction = dirlist[2] #turn
        if dH > 110 and dH <= 160: direction = dirlist[3] #sharp 
        if dH > 160 and dH <= 170: direction = dirlist[4] #uturn
        if dH > 170 and dH <= 190: direction = dirlist[5]
        if dH > 190 and dH <= 200: direction = dirlist[6] #uturn
        if dH > 200 and dH <= 250: direction = dirlist[7] #sharp
        if dH > 250 and dH <= 330: direction = dirlist[8] #turn
        if dH > 330 and dH <= 350: direction = dirlist[9] #slight
        if dH > 350 and dH <= 360: direction = dirlist[10]
    rows.append([seg[i]['wpt']['name'],seg[i]['wpt']['description'],direction,segDist,runTotal])
    
# Generate Output File Name
ext = 'csv' if args.get('csv') else 'xlsx'
if args.get('reverse'):
    filename = splitext(args.get('gpxfile'))[0] + '_reverse.'+ext
else: 
    filename = splitext(args.get('gpxfile'))[0] + '.'+ext

fldnames = ["Way Point Name","Description","Direction","Segment","Running Total"]
runTotal = 0
if args.get('csv'): # CSV Output
    logger.info('Writing data to CSV file: {}...'.format(filename))
    csvfile = open(filename,'w')
    writer = csv.writer(csvfile,delimiter=',',quoting=csv.QUOTE_ALL)
    writer.writerow(fldnames)
    for rowData in rows:
        writer.writerow(rowData)
    csvfile.close()
    logger.info('Finished writing to CSV file')
else: # Excel Output
    DEFAULT_FONT.name = 'Arial'
    logger.info('Writing data to Excel file: {}...'.format(filename))
    wb = Workbook()
    ws = wb.active
    ws.title = 'FLT Tracking'
    ws.append(fldnames)
    ws.row_dimensions[1].font = Font(name='Arial',bold=True)
    row = 2
    for rowData in rows:
        ws.append(rowData)
        row += 1
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].number_format = '##0.0'
    ws.column_dimensions['E'].number_format = '##0.0'
    #map_range = openpyxl.workbook.defined_name.DefinedName('Map_M18', attr_text='A1:E{}'.format(row-1))
    #wb.defined_names.append(map_range)
    total_row = ['Total:']
    total_row += ['' for i in range(len(fldnames)-2)]
    total_row.append('=SUM(INDEX(A:E,2,COLUMN()-1):INDEX(A:F,ROW()-1,COLUMN()-1))')
    ws.append(total_row)
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=len(fldnames)-1)
    ws.row_dimensions[row].font = Font(name='Arial',bold=True)
    ws.freeze_panes = ws['A2']
    wb.save(filename)
    logger.info('Finished writing to Excel file')
